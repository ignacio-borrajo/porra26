from datetime import timedelta
from io import BytesIO

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.tests.factories import UserFactory
from competition.tests.factories import (
    MatchFactory,
    PredictionFactory,
    RoundFactory,
    TeamFactory,
)


@pytest.mark.django_db
def test_export_requires_login(client):
    r = client.get(reverse("stats:historico_export"))
    assert r.status_code == 302


@pytest.mark.django_db
def test_export_returns_xlsx_with_headers(client):
    client.force_login(UserFactory())
    r = client.get(reverse("stats:historico_export"))
    assert r.status_code == 200
    assert r["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in r["Content-Disposition"]
    assert "historico-porra-26.xlsx" in r["Content-Disposition"]


@pytest.mark.django_db
def test_export_contents_include_header_result_and_colored_exact_cell(client):
    grp = RoundFactory(id="groups", points=3, partial_points=1, short="G", order=1)
    now = timezone.now()
    esp = TeamFactory(code="ESP", name="España", flag="🇪🇸")
    fra = TeamFactory(code="FRA", name="Francia", flag="🇫🇷")
    m = MatchFactory(
        round=grp,
        home=esp,
        away=fra,
        kickoff=now - timedelta(days=1),
        result_home=2,
        result_away=1,
        finished_at=now,
    )
    ana = UserFactory(name="Ana López", email="ana@edisa.com")
    PredictionFactory(player=ana, match=m, home=2, away=1, earned=3)

    client.force_login(ana)
    r = client.get(reverse("stats:historico_export"))
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    assert "Histórico" in wb.sheetnames
    ws = wb["Histórico"]

    header_values = [c.value for c in ws[1]]
    assert "Jugador" in header_values
    assert "ESP - FRA" in header_values
    assert "Total" in header_values

    result_row = [c.value for c in ws[2]]
    assert "2-1" in result_row

    # Tercera fila: jugador Ana López con pronóstico exacto en columna 2.
    name_cell = ws.cell(row=3, column=1)
    assert "Ana López" in (name_cell.value or "")
    exact_cell = ws.cell(row=3, column=2)
    assert exact_cell.value == "2-1"
    assert exact_cell.fill.fgColor.rgb == "FF22C55E"

    assert ws.freeze_panes == "B3"
