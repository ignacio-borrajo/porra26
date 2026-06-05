"""Serializa la matriz del Histórico a un fichero .xlsx coloreado."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stats.services.history_matrix import HistoryMatrix

_HEADER_FILL = PatternFill("solid", fgColor="FF1A1530")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_RESULT_FONT = Font(italic=True, bold=True)
_EXACT_FILL = PatternFill("solid", fgColor="FF22C55E")
_PARTIAL_FILL = PatternFill("solid", fgColor="FFF59E0B")
_HIT_FONT = Font(color="FFFFFFFF", bold=True)
_TOTAL_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def render_xlsx(matrix: HistoryMatrix) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"

    last_col = 2 + len(matrix.matches)  # A=Jugador, …matches…, última=Total

    ws.cell(row=1, column=1, value="Jugador")
    for idx, m in enumerate(matrix.matches, start=2):
        ws.cell(row=1, column=idx, value=f"{m.home_code} - {m.away_code}")
    ws.cell(row=1, column=last_col, value="Total")
    for col in range(1, last_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for idx, m in enumerate(matrix.matches, start=2):
        cell = ws.cell(row=2, column=idx, value=f"{m.result_home}-{m.result_away}")
        cell.font = _RESULT_FONT
        cell.alignment = _CENTER

    for row_idx, p in enumerate(matrix.players, start=3):
        name_cell = ws.cell(row=row_idx, column=1, value=f"{p.position}. {p.name}")
        name_cell.alignment = _LEFT
        prow = matrix.cells.get(p.id, {})
        for col_idx, m in enumerate(matrix.matches, start=2):
            cell_data = prow.get(m.id)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _CENTER
            if cell_data is None or cell_data.state == "empty":
                continue
            cell.value = f"{cell_data.home}-{cell_data.away}"
            if cell_data.state == "exact":
                cell.fill = _EXACT_FILL
                cell.font = _HIT_FONT
            elif cell_data.state == "partial":
                cell.fill = _PARTIAL_FILL
                cell.font = _HIT_FONT
        total_cell = ws.cell(row=row_idx, column=last_col, value=matrix.totals.get(p.id, 0))
        total_cell.font = _TOTAL_FONT
        total_cell.alignment = _RIGHT

    ws.freeze_panes = "B3"
    ws.column_dimensions[get_column_letter(1)].width = 28
    for col in range(2, last_col):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions[get_column_letter(last_col)].width = 10
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
