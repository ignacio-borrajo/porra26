"""Importación masiva de jugadores desde un fichero Excel (.xlsx).

Reglas: nunca actualiza usuarios existentes. Si el email ya está dado de
alta, la fila se cuenta como saltada y se continúa. No aplica validadores
de contraseña fuerte; la contraseña del Excel se asigna tal cual.
"""

from dataclasses import dataclass
from io import BytesIO

from django.core.exceptions import ValidationError
from django.core.validators import EmailValidator
from django.db import transaction

from accounts.models import AuditLog, User
from accounts.validators import validate_email_domain
from pot.models import Payment

_HEADER_ALIASES = {
    "email": "email",
    "correo": "email",
    "mail": "email",
    "nombre": "name",
    "name": "name",
    "contrasena": "password",
    "password": "password",
    "clave": "password",
    "contrasea": "password",
}

_ACCENTS = str.maketrans("áéíóúñ", "aeioun")


@dataclass
class ImportResult:
    created: int = 0
    skipped_existing: int = 0
    skipped_invalid_email: int = 0
    skipped_empty: int = 0
    error: str | None = None

    @property
    def skipped_total(self) -> int:
        return self.skipped_existing + self.skipped_invalid_email + self.skipped_empty

    @property
    def total_rows(self) -> int:
        return self.created + self.skipped_total


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().translate(_ACCENTS)


def _detect_columns(row) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = _HEADER_ALIASES.get(_normalize_header(cell))
        if key and key not in mapping:
            mapping[key] = idx
    if {"email", "name", "password"}.issubset(mapping):
        return mapping
    return None


def _cell(row, idx):
    return row[idx] if idx < len(row) else None


def import_players_from_xlsx(uploaded_file, *, actor=None) -> ImportResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ImportResult(error="Falta la dependencia openpyxl en el servidor.")

    try:
        data = uploaded_file.read() if hasattr(uploaded_file, "read") else uploaded_file
        wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return ImportResult(error="No se pudo leer el fichero. ¿Es un .xlsx válido?")

    ws = wb.active
    if ws is None:
        return ImportResult(error="El fichero no contiene hojas.")

    rows_iter = ws.iter_rows(values_only=True)
    header_map: dict[str, int] | None = None
    for row in rows_iter:
        if row is None:
            continue
        header_map = _detect_columns(row)
        if header_map is not None:
            break
    if header_map is None:
        return ImportResult(
            error="No se encontraron las columnas 'email', 'nombre' y 'contraseña'."
        )

    email_validator = EmailValidator()
    result = ImportResult()

    for row in rows_iter:
        if row is None or all(
            c is None or (isinstance(c, str) and not c.strip()) for c in row
        ):
            result.skipped_empty += 1
            continue

        raw_email = _cell(row, header_map["email"])
        raw_name = _cell(row, header_map["name"])
        raw_password = _cell(row, header_map["password"])

        email = str(raw_email).strip().lower() if raw_email is not None else ""
        name = str(raw_name).strip() if raw_name is not None else ""
        password = "" if raw_password is None else str(raw_password)

        if not email or not name or not password.strip():
            result.skipped_empty += 1
            continue

        try:
            email_validator(email)
            validate_email_domain(email)
        except ValidationError:
            result.skipped_invalid_email += 1
            continue

        if User.objects.filter(email=email).exists():
            result.skipped_existing += 1
            continue

        with transaction.atomic():
            user = User.objects.create_user(
                email=email,
                password=password,
                name=name,
                is_jugador=True,
                is_gestor=False,
                must_change_password=True,
            )
            Payment.objects.get_or_create(player=user)
            AuditLog.objects.create(
                actor=actor,
                action="player_created",
                target_type="user",
                target_id=str(user.id),
                payload={"source": "xlsx_import"},
            )
        result.created += 1

    return result
